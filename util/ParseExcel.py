# encoding = utf - 8
# 用于存放excel文件操作方法

import openpyxl
from openpyxl.styles import Border,Side,Font
from xlrd import open_workbook
import itertools
import time
#添加部分
import xlrd
from xlutils import copy
from config.VarConfig import *

class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        # 设置字体颜色
        self.font = Font(color = None)
        #颜色对应的RGB值
        self.RGBDict = {'red':'FFFF3030','green':'FF008B00'}



    def loadWorkBook(self,excelPathAndName):
        '''
        将excel文件加载到内存，并获取其workbook对象
        '''
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName

        return self.workbook


    def getSheetByName(self,sheetName):
        '''
        根据sheet名称获取sheet对象
        '''
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e


    def getSheetByIndex(self,sheetIndex):
        '''
        根据sheet索引号获取sheet对象
        '''
        try:
            sheetName = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e

        sheet = self.workbook.get_sheet_by_name(sheetName)
        return sheet


    def getRowsNumber(self,sheet):
        '''
        获取sheet中有数据区域的结束行号
        '''
        return sheet.max_row


    def getColsNumber(self,sheet):
        '''
        获取sheet中有数据区域的结束列号
        '''
        return sheet.max_column


    def getStartRowNumber(self,sheet):
        '''
        获取sheet中有数据区域的开始行号
        '''
        return sheet.min_row


    def getStartColNumber(self,sheet):
        '''
        获取sheet中有数据区域的开始列号
        '''
        return sheet.min_column


    def getRow(self,sheet,rowNo):
        '''
        获取sheet中某一行，返回该行所有数据内容组成的tuple，
        下标从1开始，sheet.rows[1]表示第一行
        '''
        try:
            myrows = sheet[rowNo]
            return myrows

            # colNo = ParseExcel().getColsNumber(sheet)
            # print("colNo:",colNo)
            # MyValue = []
            # for i in range(colNo):
            #     MyValue = MyValue.append(sheet.cell(rowNo,i).value)
            # return MyValue

        except Exception as e:
            raise e


    def getColumn(self,sheet,colNo):
        '''
        获取sheet中某一列，返回该列所有数据内容组成的tuple，
        下标从1开始，sheet.columns[1]表示第一行
        '''
        try:
            return sheet[colNo]
        except Exception as e:
            raise e


    def getCellOfValue(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        '''
        根据单元格所在的位置索引获取该单元格中的值，下表从1开始，
        sheet.cell(row = 1,column = 1).value 表示excel中第一行第一列的值
        '''
        if coordinate != None:
            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception('**********在 getCellOfValue 中无法找到该表格对象！**********')


    def getCellOfObject(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        '''
        获取某个单元格对象，可以根据单元格所在位置的数字索引
        也可以直接根据excel中单元格的编码及坐标，
        如：getCellObject(sheet,coordinate = 'A1')
        或：getCellObject(sheet,rowNo = 1,colsNo = 2)
        '''
        if coordinate != None:
            try:
                coordinate = coordinate.decode('utf-8')
                return sheet[coordinate]
                # return sheet.cell(coordinate = coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet(row = rowNo,column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception('**********在 getCellOfObject 中无法找到该表格对象！**********')


    def writeCell(self,sheet,content,coordinate = None,rowNo = None,colsNo = None,style = None):
        '''
        根据单元格在excel中的编码坐标或者数字索引坐标向单元格中写入数据，
        下表从1开始，参数style表示字体的颜色名称，如：red、green
        '''
        if coordinate is not None:
            try:
                coordinate = coordinate.decode('utf-8')
                sheet[coordinate].value = content
                # sheet.cell(coordinate = coordinate).value = content
                if style is not None:
                    sheet[coordinate].font = Font(color = self.RGBDict[style])
                    # sheet.cell(coordinate = coordinate).font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                # sheet.cell(row = rowNo,column = colsNo).value = ""
                sheet.cell(row = rowNo,column = colsNo).value = content
                if style:
                    sheet.cell(row = rowNo,column = colsNo).font = Font(color = self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception('**********在 writeCell 中无法找到该表格对象！**********')


    def writeCellCurrentTime(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        '''
        写入当前时间，下表从1开始
        '''
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                # sheet.cell(coordinate = coordinate).value = currentTime
                coordinate = coordinate.decode('utf-8')
                sheet[coordinate].value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception('**********在 writeCellCurrentTime 中无法找到该表格对象！**********')


    def randomContracNum(self,myInfo):      # 写值失败，暂时弃用
        try:
            from action.PageAction import randomNum
            splitInfo = myInfo.split("|")
            sheetName = splitInfo[0]
            myCol = int(splitInfo[1])
            myStr = splitInfo[2] + randomNum(8)
            print("********** 随机字符串为：", myStr, " **********")

            excelObj = ParseExcel()
            excelObj.loadWorkBook(dataFilePath)
            sheet = excelObj.getSheetByName(sheetName)

            isExecuteColumn = excelObj.getColumn(sheet, DataSource_IsExecute)

            for ltime, i in enumerate(isExecuteColumn[1:]):
                if i.value == "未使用":
                    # coordinate = myCol + str(ltime+2)
                    # coordinate = str(coordinate).encode('utf-8')
                    # # sheet[coordinate] = myStr
                    # # excelObj.writeCell(sheet, myStr, rowNo=ltime+2, colsNo=myCol)
                    # excelObj.writeCellCurrentTime(sheet,coordinate=coordinate)

                    # print("前：",sheet.cell(ltime+2,myCol).value)
                    # sheet.cell(ltime + 2, myCol).value = myStr
                    # sheet["L2"].value = myStr
                    # sheet["L2"] = myStr
                    # excelObj.writeCell(sheet,myStr,rowNo=ltime+2,colsNo=myCol)
                    # excelObj.writeCellCurrentTime(sheet, rowNo=ltime+2, colsNo=3)
                    # print("后：",sheet.cell(ltime+2,myCol).value)
                    # excelObj.workbook.save(dataFilePath)

                    print("前：", sheet.cell(ltime + 2, myCol).value)
                    excelObj.writeCell(sheet,myStr,rowNo=ltime+2,colsNo=myCol)
                    excelObj.writeCellCurrentTime(sheet,rowNo=ltime+2,colsNo=3)
                    time.sleep(2)

                    print("后：", sheet.cell(ltime + 2, myCol).value)

                    break
        except Exception as e:
            raise e


# *************************************************优化后内容*************************************************

class ParseExcel_new(object):

    def __init__(self, file_name=None, id=None):
        if file_name:
            self.file_name = file_name
            self.id = id
            self.data = self.get_data()
        else:
            self.file_name = '../dataconfig/interface.xls'
            self.id = 0
            self.data = self.get_data()

    def get_data(self):
        data = xlrd.open_workbook(self.file_name)
        tables = data.sheets()[self.id]
        return tables

    def get_lines(self):
        tables = self.data
        return tables.nrows

    def get_cell_value(self,row,col):
        return self.data.cell_value(row,col)

    def write_value(self,row, col, value):
        read_data = xlrd.open_workbook(self.file_name)
        write_data = copy.copy(read_data)
        sheet_data = write_data.get_sheet(0)
        sheet_data.write(row,col,value)
        write_data.save(self.file_name)

    #根据对应的caseid找到对应行的内容
    def get_rows_data(self,case_id):
        row_num = self.get_row_num(case_id)
        row_data = self.get_row_values(row_num)
        return row_data

    #根据对应的caseid找到对应的行号
    def get_row_num(self,case_id):
        num = 0
        cols_data = self.get_col_values()
        for col_data in cols_data:
            if case_id in col_data:
                return num
            num = num + 1
        return num

    #根据行号，找到该行的内容
    def get_row_values(self,row):
        tables = self.data
        row_data = tables.row_values(row)
        return row_data

    #获取某一列的内容
    def get_col_values(self,col_id=None):
        tables = self.data
        if col_id != None:
            col_data = tables.col_values(col_id)
        else:
            col_data = tables.col_values(0)
        return col_data

if __name__ == "__main__":
    pe = ParseExcel()
    #调用excel数据，表格可替换
    pe.loadWorkBook(u'E:\\python相关\\工程\\销售合同新增\\testData\\数据汇总.xlsx')
    # print("通过名称获取sheet对象的名字：",pe.getSheetByName(u'Contract_01').title)
    print("通过index序号获取sheet对象的名字：",pe.getSheetByName("项目合同_数据").title)

    # sheet = pe.getSheetByIndex(0)
    sheet = pe.getSheetByName("项目合同_数据")
    print(type(sheet))
    #获取有数据区域的最大行列号
    print("最大行号：",pe.getRowsNumber(sheet))
    print("最大列号：",pe.getColsNumber(sheet))

    #获取第一行,getRow和getColumn迭代报错，需调试
    print('第3行第12列:',sheet.cell(3,12))
    rows = pe.getRow(sheet,3)
    for i in rows:
        print('第',i,'行值为：','\n',i.value)

    # 获取某一列所有数据
    columns = pe.getColumn(sheet,'E')
    for j in columns:
        print('第',j,'列值为：','\n',j.value)


    #获取第一行第一列单元格内容
    print('第一行第一列单元格内容：',pe.getCellOfValue(sheet,rowNo = 1,colsNo = 1))

    '''
    #写操作调试内容，后期可做处理
    pe.writeCell(sheet,u"成功",rowNo = 2,colsNo = 12)
    # sheet.cell(row=5, column=5).value = "chenggong"
    print(sheet.cell(row=5,column=5).value)

    pe.writeCellCurrentTime(sheet,rowNo = 5,colsNo = 6)
    '''

    pe.randomContracNum("项目合同_数据|12|销售合同")

    # a = str("L7").encode('utf-8')
    #
    # print("单元格为：",a)
    # # print(a,"单元格为：",sheet[a])
    # print("1_L7值为：",sheet[a].value)
    # sheet[a] = "123123"
    # print("2_L7值为：",sheet[a].value)


    # print("1_G7值为：",sheet["G7"].value)
    # pe.writeCell(sheet, content="135246", rowNo=7, colsNo=7)
    # print("2_G7值为：",sheet["G7"].value)

    # print('000000000000000',sheet.cell(coordinate="E2").value)

    print('**********ParseExcel处理成功**********')